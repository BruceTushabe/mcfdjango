
from core.models import Account  # Import your Account model
from openpyxl import load_workbook

def find_account(account_number):
    """Find an account by account number in the Excel sheet."""
    try:
        account_number = int(account_number)  # Convert to integer for comparison
        workbook = load_workbook('data1.xlsx', data_only=True)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[2] == account_number:
                return row  # Return the row data if account number is found
    except Exception as e:
        print(f"Error finding account data: {e}")
    return None  # Return None if account number is not found


"""def find_account(account_number):

    try:
        # Assuming Account model has a field 'account_number'
       
        account = Account.objects.get(account_number=account_number)
        return account  # Return the found account object
    except Account.DoesNotExist:
        return None  # Account with the given account number does not exist
"""

def populate_excel(account_number, form_data):
    """Populate"""
    # Load the Excel workbook
    workbook = load_workbook('data1.xlsx')

    # Select the active worksheet
    worksheet = workbook.active

    # Find the row corresponding to the given account number
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[0] == account_number:
            # Fill in the form data to the respective columns
            worksheet.cell(row=row[23], column=24).value = form_data['loan_application_date']
            worksheet.cell(row=row[24], column=25).value = form_data['purpose_of_the_loan']
            worksheet.cell(row=row[25], column=26).value = form_data['address_location']
            worksheet.cell(row=row[26], column=27).value = form_data['business_financed']
            worksheet.cell(row=row[27], column=28).value = form_data['group_name']
            worksheet.cell(row=row[28], column=29).value = form_data['reason_for_default_summarised']
            worksheet.cell(row=row[29], column=30).value = form_data['detailed_reason_for_default']

            # Save the workbook
            workbook.save('data1.xlsx')
            return

    # If the account number is not found, you can handle this case accordingly
    raise ValueError("Account number not found in the Excel sheet")